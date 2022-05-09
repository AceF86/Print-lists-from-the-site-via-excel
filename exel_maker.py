import json

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Side, Font, DEFAULT_FONT
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.dimensions import RowDimension


def create_exel(time, judg, secretary, json_data):

    book = openpyxl.Workbook()
    sheet = book.active
    RowDimension(sheet, bestFit=True)

    alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    alignment2 = Alignment(wrap_text=True, horizontal="right")
    thins = Side(border_style="thin")
    double = Side(border_style="thin")

    font = Font(bold=True, name="Times New Roman")
    font2 = Font(size=16, name="Times New Roman")
    font3 = Font(size=15, name="Times New Roman")

    den = time
    a = judg
    s = secretary

    sheet.merge_cells("A1:D1")
    sheet.merge_cells("A2:D2")
    sheet["A1"].value = f"СПИСОК СПРАВ, ПРИЗНАЧЕНИХ ДО РОЗГЛЯДУ на {den} рік"
    sheet["A2"].value = f"Суддя: {a}"
    sheet["A3"] = "Час"
    sheet["B3"] = "Номер справи"
    sheet["C3"] = "Сторони по справі"
    sheet["D3"] = "Суть справи"

    sheet["A1"].alignment = alignment
    sheet["A2"].alignment = alignment
    sheet["A3"].alignment = alignment
    sheet["B3"].alignment = alignment
    sheet["C3"].alignment = alignment
    sheet["D3"].alignment = alignment

    sheet["A1"].font = font2
    sheet["A2"].font = font3
    sheet["A3"].font = font
    sheet["B3"].font = font
    sheet["C3"].font = font
    sheet["D3"].font = font

    try:
        url = "https://pr.zk.court.gov.ua/new.php"

        payload = "q_court_id=0708"
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://pr.zk.court.gov.ua/sud0708/gromadyanam/csz/",
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        data = response.json()
    except Exception as ex:
        with open(json_data, "r", encoding="utf-8") as f:
            file_content = f.read()
            data = json.loads(file_content)

    row = 4
    for i in data:
        if den in i["date"] and a in i["judge"]:
            sheet[row][0].value = i["date"][10:]
            sheet[row][1].value = i["number"]
            sheet[row][2].value = "\n" + i["involved"] + "\n"
            sheet[row][3].value = "\n" + i["description"] + "\n"
            row += 1

    for row in sheet.iter_rows(3):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                shrink_to_fit=False,
                vertical="center",
                horizontal="center",
            )
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)
            cell.font = Font(size=13, name="Times New Roman")

    sheet.column_dimensions["A"].width = 9
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 51
    sheet.column_dimensions["D"].width = 51

    sheet.row_dimensions[2].height = 25
    sheet.row_dimensions[3].height = 35

    newRowLocation = sheet.max_row + 3
    sheet.cell(
        column=1, row=newRowLocation, value="Секретар судового засідання"
    ).font = Font(name="Times New Roman")
    sheet.cell(column=4, row=newRowLocation, value=s).alignment = alignment2
    sheet.cell(column=4, row=newRowLocation, value=s).font = Font(
        name="Times New Roman"
    )

    DEFAULT_FONT.size = "13"
    sheet.print_options.horizontalCentered = True
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.fitToPage = True
    sheet.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.200, bottom=0.1, header=0.3, footer=0.3
    )

    book.save("book.xlsx")
    book.close()
